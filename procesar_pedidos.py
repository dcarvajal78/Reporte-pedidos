"""
╔══════════════════════════════════════════════════════════════════╗
║         PROCESADOR DIARIO DE PEDIDOS — AUTOPLANET               ║
║  Uso: python procesar_pedidos.py ARCHIVO.xls                    ║
║  Genera: Pedidos_Reporte_YYYY-MM-DD.xlsx                        ║
╚══════════════════════════════════════════════════════════════════╝
"""

import sys
import os
import pandas as pd
import io
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.formatting.rule import ColorScaleRule

# ══════════════════════════════════════════════════════════════════
# CONFIGURACIÓN — ajusta estos valores si cambia algo
# ══════════════════════════════════════════════════════════════════
CONFIG = {
    "encoding":        "utf-16",          # encoding del archivo fuente
    "separador":       "\t",              # separador de columnas
    "fila_header":     3,                 # línea con los nombres de columna (0-indexed)
    "top_clientes":    100,                # cuántos clientes mostrar
    "top_skus":        100,                # cuántos SKUs mostrar
    "top_localidades": 20,                # cuántas localidades mostrar
    "top_vendedores":  50,                # cuántos vendedores mostrar
    "tipo_nc":         "ZDEV",            # ClDocVenta que = nota de crédito
    "col_canal":       "Den.Of.Vta",
    "col_cliente":     "Nmbr Solic",
    "col_material":    "Texto breve material",
    "col_fecha":       "FchCrea.Pe",
    "col_vendedor":    "Vendedor",
    "col_nombre_vend": "Nombre Ven",
    "col_venta":       "Mon.Sol.$",
    "col_factura":     "Mon.Fac.$",
    "col_rechaz":      "Mnt.Rechaz",
    "col_cant":        "Cant. Ori.",
    "col_entrega":     "Entrega",
    "col_doc":         "Doc.vta.",
    "col_sku":         "SKU SAP",
    "col_estatus_sm":  "Estatus SM",
    "col_estatus_fa":  "Estatus FA",
    "col_tipo_doc":    "ClDocVenta",
    "col_bloq":        "BloqEntreg",
    "col_den_bloq":    "Den.Bl.Ent",
    "col_mot_rech":    "Mot. Rech.",
}

# ══════════════════════════════════════════════════════════════════
# COLORES
# ══════════════════════════════════════════════════════════════════
C = {
    "NAVY":   "1B3A6B", "BLUE":   "2563EB", "LBLUE":  "DBEAFE",
    "GREEN":  "16A34A", "LGREEN": "DCFCE7", "AMBER":  "D97706",
    "LAMBER": "FEF3C7", "RED":    "DC2626", "LRED":   "FEE2E2",
    "GRAY":   "F1F5F9", "WHITE":  "FFFFFF", "LGRAY2": "E2E8F0",
    "ORANGE": "EA580C", "LORANG": "FFEDD5", "DGRAY":  "64748B",
}

def hfill(c): return PatternFill("solid", fgColor=c)
def hfont(bold=True, color="FFFFFF", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)
def bdr():
    s = Side(style='thin', color='CBD5E1')
    return Border(left=s, right=s, top=s, bottom=s)
def hdr(ws, row, cols, labels, fill=C["NAVY"]):
    for col, label in zip(cols, labels):
        c = ws.cell(row=row, column=col, value=label)
        c.font = hfont(); c.fill = hfill(fill)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = bdr()
def drow(ws, row, cs, vals, alt=False):
    bg = C["GRAY"] if alt else C["WHITE"]
    for i, v in enumerate(vals):
        c = ws.cell(row=row, column=cs+i, value=v)
        c.fill = hfill(bg); c.font = Font(name="Arial", size=10); c.border = bdr()
        c.alignment = Alignment(
            horizontal='right' if isinstance(v, (int, float)) and not isinstance(v, bool) else 'left',
            vertical='center')
def title_row(ws, text, mr, rh=28):
    ws.merge_cells(mr); sc = mr.split(':')[0]; c = ws[sc]
    c.value = f"  {text}"; c.font = Font(name="Arial", bold=True, size=13, color=C["WHITE"])
    c.fill = hfill(C["NAVY"]); c.alignment = Alignment(horizontal='left', vertical='center')
    rn = int(''.join(filter(str.isdigit, sc))); ws.row_dimensions[rn].height = rh
def total_row(ws, r, n_cols, value_cols={}):
    for col in range(1, n_cols+1):
        ws.cell(r, col).fill = hfill(C["NAVY"])
        ws.cell(r, col).font = Font(name="Arial", bold=True, color=C["WHITE"]); ws.cell(r, col).border = bdr()
    ws.cell(r, 1).value = "TOTAL"; ws.cell(r, 1).alignment = Alignment(horizontal='center', vertical='center')
    for col, val in value_cols.items():
        ws.cell(r, col).value = val
        ws.cell(r, col).alignment = Alignment(horizontal='right', vertical='center')

# ══════════════════════════════════════════════════════════════════
# 1. LEER Y LIMPIAR DATOS
# ══════════════════════════════════════════════════════════════════
def leer_archivo(ruta):
    print(f"📂 Leyendo: {ruta}")
    with open(ruta, 'r', encoding=CONFIG["encoding"]) as f:
        lines = f.readlines()

    fecha_archivo = lines[0].strip()
    hdr_idx = CONFIG["fila_header"]
    data_lines = [lines[hdr_idx]] + [l for l in lines[hdr_idx+2:] if l.strip()]
    df = pd.read_csv(io.StringIO(''.join(data_lines)), sep=CONFIG["separador"], dtype=str)
    df.columns = [c.strip() for c in df.columns]
    df = df.drop(columns=['Unnamed: 0'], errors='ignore')
    return df, fecha_archivo

def limpiar(df):
    def cn(s):
        if pd.isna(s): return np.nan
        try: return float(str(s).strip().replace('.','').replace(',','.'))
        except: return np.nan

    num_cols = [CONFIG["col_venta"], CONFIG["col_factura"], CONFIG["col_rechaz"],
                CONFIG["col_cant"], "CtdEntrega", "Neto", "Bruto", "Mnt.bloque"]
    for col in num_cols:
        if col in df.columns: df[col] = df[col].apply(cn)

    date_cols = [CONFIG["col_fecha"], "Fch.Cre.En", "Fe.SM real", "FechaFact."]
    for col in date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], format='%d.%m.%Y', errors='coerce')

    str_cols = [CONFIG["col_doc"], CONFIG["col_sku"], CONFIG["col_vendedor"],
                CONFIG["col_entrega"], CONFIG["col_tipo_doc"]]
    for col in str_cols:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: str(x).strip() if pd.notna(x) else x)

    df['Canal']       = df[CONFIG["col_canal"]].str.strip()
    df['Cliente']     = df[CONFIG["col_cliente"]].str.strip()
    df['Descripcion'] = df[CONFIG["col_material"]].str.strip()
    df['Fecha_str']   = df[CONFIG["col_fecha"]].dt.strftime('%d/%m/%Y')
    df['Estatus_SM']  = df[CONFIG["col_estatus_sm"]].str.strip()
    df['Estatus_FA']  = df[CONFIG["col_estatus_fa"]].str.strip()
    df['Es_ZDEV']     = df[CONFIG["col_tipo_doc"]] == CONFIG["tipo_nc"]
    df['Rechaz']      = df[CONFIG["col_rechaz"]].fillna(0)
    df['Sin_Entrega'] = df[CONFIG["col_entrega"]].isna() | (df['Estatus_SM'] == 'A')
    df['Fac_Neto']    = np.where(df['Es_ZDEV'], -df[CONFIG["col_factura"]], df[CONFIG["col_factura"]])
    return df

# ══════════════════════════════════════════════════════════════════
# 2. CALCULAR PIVOTS
# ══════════════════════════════════════════════════════════════════
def calcular_pivots(df):
    base = df[~df['Es_ZDEV']]
    nc   = df[df['Es_ZDEV']]
    sin_e= base[base['Sin_Entrega']]

    # KPIs globales
    kpis = {
        "total_pedidos":   base[CONFIG["col_doc"]].nunique(),
        "total_lineas":    len(base),
        "total_venta":     base[CONFIG["col_venta"]].sum(),
        "total_fac_bruto": base[CONFIG["col_factura"]].sum(),
        "total_nc":        nc[CONFIG["col_factura"]].sum(),
        "total_rechaz":    base['Rechaz'].sum(),
        "total_se_ped":    sin_e[CONFIG["col_doc"]].nunique(),
        "total_se_val":    sin_e[CONFIG["col_venta"]].sum(),
    }
    kpis["total_fac_neto"] = kpis["total_fac_bruto"] - kpis["total_nc"]
    ped_sm = base[base['Estatus_SM']=='C'][CONFIG["col_doc"]].nunique()
    kpis["tasa_sm"] = ped_sm / kpis["total_pedidos"] * 100 if kpis["total_pedidos"] else 0

    # Canal
    canal_df = base.groupby('Canal').agg(
        Pedidos=   (CONFIG["col_doc"],  'nunique'),
        Lineas=    (CONFIG["col_doc"],  'count'),
        Unidades=  (CONFIG["col_cant"], 'sum'),
        Venta_Sol= (CONFIG["col_venta"],'sum'),
        Fac_Bruto= (CONFIG["col_factura"],'sum'),
        Rechazado= ('Rechaz','sum'),
    ).reset_index()
    nc_canal = nc.groupby('Canal')[CONFIG["col_factura"]].sum().reset_index().rename(columns={CONFIG["col_factura"]:'NC'})
    canal_df = canal_df.merge(nc_canal, on='Canal', how='left').fillna({'NC':0})
    canal_df['Fac_Neto'] = canal_df['Fac_Bruto'] - canal_df['NC']
    canal_df['Part_%']   = canal_df['Venta_Sol'] / canal_df['Venta_Sol'].sum()
    canal_df = canal_df.sort_values('Venta_Sol', ascending=False)

    # Cliente
    # Vendedor principal por cliente (el de mayor venta)
    vend_principal = base.groupby(['Canal','Cliente', CONFIG["col_nombre_vend"]])[ CONFIG["col_venta"]].sum().reset_index()
    vend_principal = vend_principal.sort_values(CONFIG["col_venta"], ascending=False).drop_duplicates(['Canal','Cliente'])[['Canal','Cliente', CONFIG["col_nombre_vend"]]]

    cliente_df = base.groupby(['Canal','Cliente']).agg(
        Pedidos=  (CONFIG["col_doc"],    'nunique'),
        Lineas=   (CONFIG["col_doc"],    'count'),
        Venta_Sol=(CONFIG["col_venta"],  'sum'),
        Fac_Bruto=(CONFIG["col_factura"],'sum'),
        Rechazado=('Rechaz','sum'),
    ).reset_index()
    cliente_df = cliente_df.merge(vend_principal, on=['Canal','Cliente'], how='left')
    nc_cli = nc.groupby(['Canal','Cliente'])[CONFIG["col_factura"]].sum().reset_index().rename(columns={CONFIG["col_factura"]:'NC'})
    cliente_df = cliente_df.merge(nc_cli, on=['Canal','Cliente'], how='left').fillna({'NC':0})
    cliente_df['Fac_Neto'] = cliente_df['Fac_Bruto'] - cliente_df['NC']
    cliente_df = cliente_df.sort_values('Venta_Sol', ascending=False).head(CONFIG["top_clientes"])

    # Fecha
    fecha_df = base.groupby(['Canal','Fecha_str']).agg(
        Pedidos=  (CONFIG["col_doc"],  'nunique'),
        Lineas=   (CONFIG["col_doc"],  'count'),
        Venta_Sol=(CONFIG["col_venta"],'sum'),
        Rechazado=('Rechaz','sum'),
    ).reset_index().sort_values('Fecha_str')

    # SKU
    sku_df = base.rename(columns={CONFIG["col_sku"]: "SKU_SAP"}).groupby(["Canal","SKU_SAP","Descripcion"]).agg(
        Unidades= (CONFIG["col_cant"],'sum'),
        Lineas=   (CONFIG["col_doc"],'count'),
        Venta_Sol=(CONFIG["col_venta"],'sum'),
        Rechazado=('Rechaz','sum'),
        Clientes= ('Cliente','nunique'),
    ).reset_index().sort_values('Venta_Sol', ascending=False).head(CONFIG["top_skus"])

    # Localidad
    loc_df = base.groupby(['Canal','Localidad']).agg(
        Pedidos=  (CONFIG["col_doc"],  'nunique'),
        Venta_Sol=(CONFIG["col_venta"],'sum'),
        Clientes= ('Cliente','nunique'),
    ).reset_index().sort_values('Venta_Sol', ascending=False).head(CONFIG["top_localidades"])

    # Vendedores
    vend_df = base.groupby(['Canal', CONFIG["col_vendedor"], CONFIG["col_nombre_vend"]]).agg(
        Pedidos=  (CONFIG["col_doc"],    'nunique'),
        Lineas=   (CONFIG["col_doc"],    'count'),
        Clientes= ('Cliente',            'nunique'),
        Venta_Sol=(CONFIG["col_venta"],  'sum'),
        Fac_Bruto=(CONFIG["col_factura"],'sum'),
        Rechazado=('Rechaz',             'sum'),
    ).reset_index()
    nc_vend = nc.groupby([CONFIG["col_vendedor"], CONFIG["col_nombre_vend"]])[CONFIG["col_factura"]].sum().reset_index().rename(columns={CONFIG["col_factura"]:'NC_ZDEV'})
    se_vend = sin_e.groupby([CONFIG["col_vendedor"], CONFIG["col_nombre_vend"]]).agg(
        SE_Ped=(CONFIG["col_doc"],  'nunique'),
        SE_Val=(CONFIG["col_venta"],'sum'),
    ).reset_index()
    vend_df = vend_df.merge(nc_vend, on=[CONFIG["col_vendedor"], CONFIG["col_nombre_vend"]], how='left').fillna({'NC_ZDEV':0})
    vend_df = vend_df.merge(se_vend, on=[CONFIG["col_vendedor"], CONFIG["col_nombre_vend"]], how='left').fillna({'SE_Ped':0,'SE_Val':0})
    vend_df['Fac_Neto'] = vend_df['Fac_Bruto'] - vend_df['NC_ZDEV']
    vend_df = vend_df.sort_values('Venta_Sol', ascending=False).head(CONFIG["top_vendedores"])

    # Sin entrega resumen
    se_res = sin_e.groupby(['Canal','Cliente', CONFIG["col_vendedor"], CONFIG["col_nombre_vend"]]).agg(
        Pedidos=    (CONFIG["col_doc"],  'nunique'),
        Lineas=     (CONFIG["col_doc"],  'count'),
        Venta_Sol=  (CONFIG["col_venta"],'sum'),
        Bloq=       (CONFIG["col_bloq"], lambda x: (x.notna()&(x.str.strip()!='')).sum()),
        Motivo=     (CONFIG["col_den_bloq"], lambda x: x.dropna().replace('',np.nan).dropna().mode()[0] if len(x.dropna().replace('',np.nan).dropna())>0 else ''),
    ).reset_index().sort_values('Venta_Sol', ascending=False)
    # Build doc+fecha pairs per client
    doc_fecha = sin_e.drop_duplicates(subset=[CONFIG["col_doc"]]).sort_values(CONFIG["col_fecha"])
    doc_fecha_grp = doc_fecha.groupby(['Canal','Cliente', CONFIG["col_vendedor"], CONFIG["col_nombre_vend"]]).apply(
        lambda x: [(str(r[CONFIG["col_doc"]]), r[CONFIG["col_fecha"]].strftime('%d/%m/%Y') if pd.notna(r[CONFIG["col_fecha"]]) else '') 
                   for _, r in x.iterrows()]
    ).reset_index().rename(columns={0:'DocFechaPairs'})
    se_res = se_res.merge(doc_fecha_grp, on=['Canal','Cliente', CONFIG["col_vendedor"], CONFIG["col_nombre_vend"]], how='left')


    # NC detalle
    nc_det = nc[['Canal','Cliente',CONFIG["col_doc"],'Fecha_str',CONFIG["col_sku"],'Descripcion',
                 CONFIG["col_cant"],CONFIG["col_factura"],CONFIG["col_vendedor"],CONFIG["col_nombre_vend"]]].copy()

    # Sin entrega detalle
    se_det = sin_e[['Canal',CONFIG["col_doc"],'Cliente','Localidad','Fecha_str',
                    CONFIG["col_sku"],'Descripcion',CONFIG["col_cant"],CONFIG["col_venta"],
                    'Estatus_SM',CONFIG["col_bloq"],CONFIG["col_den_bloq"],
                    CONFIG["col_vendedor"],CONFIG["col_nombre_vend"]]].copy()


    # ── Con Entrega y No Facturado ─────────────────────────────────────────
    base['Tiene_Entrega'] = base[CONFIG["col_entrega"]].notna() & (base[CONFIG["col_entrega"]].str.strip() != '') & (base[CONFIG["col_entrega"]].str.strip() != 'nan')
    base['No_Facturado']  = base[CONFIG["col_factura"]].fillna(0) == 0
    cenf = base[base['Tiene_Entrega'] & base['No_Facturado']].copy()

    # Resumen por cliente
    cenf_res = cenf.groupby(['Canal','Cliente', CONFIG["col_entrega"], CONFIG["col_doc"], CONFIG["col_vendedor"], CONFIG["col_nombre_vend"]]).agg(
        Lineas    =(CONFIG["col_doc"],    'count'),
        Venta_Sol =(CONFIG["col_venta"],  'sum'),
        Fac       =(CONFIG["col_factura"],'sum'),
        Rechazado =('Rechaz',             'sum'),
    ).reset_index()
    cenf_res['Diferencia'] = cenf_res['Venta_Sol'] - cenf_res['Fac']
    cenf_res = cenf_res.sort_values('Venta_Sol', ascending=False)

    # KPIs
    kpis['cenf_docs']   = cenf[CONFIG["col_doc"]].nunique()
    kpis['cenf_venta']  = cenf[CONFIG["col_venta"]].sum()
    kpis['cenf_canales']= cenf['Canal'].nunique()

    # Detalle lineas
    cenf_det = cenf[['Canal', CONFIG["col_doc"], CONFIG["col_entrega"], 'Cliente', 'Localidad', 'Fecha_str',
                      CONFIG["col_sku"], 'Descripcion', CONFIG["col_cant"],
                      CONFIG["col_venta"], CONFIG["col_factura"], 'Rechaz',
                      CONFIG["col_vendedor"], CONFIG["col_nombre_vend"]]].copy()

    return kpis, canal_df, cliente_df, fecha_df, sku_df, loc_df, vend_df, se_res, se_det, nc_det, base, sin_e, cenf_res, cenf_det

# ══════════════════════════════════════════════════════════════════
# 3. ESCRIBIR EXCEL
# ══════════════════════════════════════════════════════════════════
def escribir_excel(df, fecha_archivo, kpis, canal_df, cliente_df, fecha_df,
                   sku_df, loc_df, vend_df, se_res, se_det, nc_det, base, ruta_salida,
                   cenf_res=None, cenf_det=None):

    tv = kpis["total_venta"]; tfb = kpis["total_fac_bruto"]
    tnc = kpis["total_nc"];   tfn = kpis["total_fac_neto"]
    tr  = kpis["total_rechaz"]

    # ── Datos Limpios (pandas, rápido) ──────────────────────────────────────
    keep = [CONFIG["col_doc"],'Posición',CONFIG["col_entrega"],CONFIG["col_tipo_doc"],
            'Centro','Solicitan.',CONFIG["col_cliente"],CONFIG["col_canal"],
            'Denominación','Nombre Destinatario','Región','Localidad',
            CONFIG["col_fecha"],'BloqEntreg','Den.Bl.Ent','Moneda',
            'MotivRech',CONFIG["col_mot_rech"],CONFIG["col_rechaz"],
            'Material WMS',CONFIG["col_sku"],CONFIG["col_material"],
            CONFIG["col_cant"],'CtdEntrega','UM venta',
            CONFIG["col_venta"],CONFIG["col_factura"],'Fac_Neto','Factura',
            'Fe.SM real',CONFIG["col_estatus_sm"],'FechaFact.',CONFIG["col_estatus_fa"],
            'E.Transpor',CONFIG["col_vendedor"],CONFIG["col_nombre_vend"]]
    keep = [c for c in keep if c in df.columns]
    df_clean = df[keep].copy()
    for col in [CONFIG["col_fecha"], 'Fe.SM real', 'FechaFact.']:
        if col in df_clean.columns:
            df_clean[col] = df_clean[col].dt.strftime('%d/%m/%Y').fillna('')

    with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
        df_clean.to_excel(writer, sheet_name='📋 Datos Limpios', index=False, startrow=1)

    wb = load_workbook(ruta_salida)
    ws_d = wb['📋 Datos Limpios']
    ws_d.sheet_view.showGridLines = False; ws_d.sheet_properties.tabColor = C["BLUE"]
    nc2 = ws_d.max_column
    ws_d.insert_rows(1)
    ws_d.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc2)
    c = ws_d['A1']
    c.value = f"  DATOS LIMPIOS — {fecha_archivo}  |  Fac_Neto = Mon.Fac.$ − Notas de Crédito ZDEV"
    c.font = Font(name="Arial", bold=True, size=10, color=C["WHITE"])
    c.fill = hfill(C["NAVY"]); c.alignment = Alignment(horizontal='left', vertical='center')
    ws_d.row_dimensions[1].height = 22
    for col in range(1, nc2+1):
        c = ws_d.cell(2, col); c.font = hfont(); c.fill = hfill(C["NAVY"])
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); c.border = bdr()
    ws_d.row_dimensions[2].height = 28
    ws_d.freeze_panes = 'A3'
    ws_d.auto_filter.ref = f"A2:{get_column_letter(nc2)}{ws_d.max_row}"

    # ── Dashboard ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("📊 Dashboard", 0)
    ws.sheet_view.showGridLines = False; ws.sheet_properties.tabColor = C["NAVY"]
    ws.merge_cells('A1:M1')
    c = ws['A1']; c.value = f"  REPORTE DE PEDIDOS Y VENTAS  |  {fecha_archivo}"
    c.font = Font(name="Arial", bold=True, size=16, color=C["WHITE"]); c.fill = hfill(C["NAVY"])
    c.alignment = Alignment(horizontal='left', vertical='center'); ws.row_dimensions[1].height = 36
    ws.merge_cells('A2:M2'); ws['A2'].fill = hfill(C["BLUE"]); ws.row_dimensions[2].height = 5

    kpi_list = [
        ("PEDIDOS TOTALES",    f"{kpis['total_pedidos']:,.0f}",       C["NAVY"],   C["LBLUE"]),
        ("VENTA SOLICITADA",   f"${tv/1e6:.1f}M",                     C["BLUE"],   C["LBLUE"]),
        ("FAC. BRUTO",         f"${tfb/1e6:.1f}M",                    C["GREEN"],  C["LGREEN"]),
        ("NOTAS CRÉD. ZDEV",   f"-${tnc/1e6:.1f}M",                   C["RED"],    C["LRED"]),
        ("FAC. NETO",          f"${tfn/1e6:.1f}M",                    C["GREEN"],  C["LGREEN"]),
        ("RECHAZADO",          f"${tr/1e6:.1f}M",                     C["ORANGE"], C["LORANG"]),
        ("TASA ENTREGA WMS",   f"{kpis['tasa_sm']:.1f}%",
            C["GREEN"] if kpis['tasa_sm']>90 else C["AMBER"],
            C["LGREEN"] if kpis['tasa_sm']>90 else C["LAMBER"]),
    ]
    for i, (title, value, color, bg) in enumerate(kpi_list):
        col = i*2+1
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
        c = ws.cell(3, col, title); c.font = Font(name="Arial", bold=True, size=8, color=color)
        c.fill = hfill(bg); c.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col+1)
        c = ws.cell(4, col, value); c.font = Font(name="Arial", bold=True, size=15, color=color)
        c.fill = hfill(bg); c.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+1)
        ws.cell(6, col).fill = hfill(color)
    for r in [3,4,5,6,7]: ws.row_dimensions[r].height = 18
    ws.row_dimensions[4].height = 28

    ws.merge_cells('A8:I8'); c = ws['A8']
    c.value = "  RESUMEN POR CANAL  (Fac. Neto = Fac. Bruto − NC ZDEV)"
    c.font = Font(name="Arial", bold=True, size=11, color=C["WHITE"])
    c.fill = hfill(C["NAVY"]); c.alignment = Alignment(horizontal='left', vertical='center'); ws.row_dimensions[8].height = 22
    hdr(ws, 9, range(1,10), ["Canal","Pedidos","Líneas","Unidades","Venta Sol. ($)","Fac. Bruto ($)","NC ZDEV ($)","Fac. Neto ($)","Rechazado ($)"], C["BLUE"])
    for i, row in canal_df.iterrows():
        r = 10 + list(canal_df.index).index(i); nc_val = row.get('NC', 0)
        vals = [row['Canal'],int(row['Pedidos']),int(row['Lineas']),int(row['Unidades']),
                row['Venta_Sol'],row['Fac_Bruto'],nc_val,row['Fac_Bruto']-nc_val,row['Rechazado']]
        drow(ws, r, 1, vals, alt=(r%2==0))
        for col in [5,6,7,8,9]: ws.cell(r,col).number_format='#,##0'
        if nc_val>0: ws.cell(r,7).fill=hfill(C["LRED"]); ws.cell(r,7).font=Font(name="Arial",color=C["RED"])
        if row['Rechazado']>0: ws.cell(r,9).fill=hfill(C["LORANG"])
    tr_r = 10+len(canal_df)
    total_row(ws, tr_r, 9, value_cols={2:kpis['total_pedidos'],3:kpis['total_lineas'],
        5:tv,6:tfb,7:tnc,8:tfn,9:tr})
    for col in [5,6,7,8,9]: ws.cell(tr_r,col).number_format='#,##0'

    ws.merge_cells('K8:Q8'); c=ws['K8']
    c.value="  TOP 10 CLIENTES"; c.font=Font(name="Arial",bold=True,size=11,color=C["WHITE"])
    c.fill=hfill(C["NAVY"]); c.alignment=Alignment(horizontal='left',vertical='center')
    hdr(ws,9,range(11,18),["Canal","Cliente","Pedidos","Venta Sol. ($)","Fac. Neto ($)","Rechazado ($)","Part. %"],C["BLUE"])
    for j,row in enumerate(cliente_df.head(10).itertuples(index=False),1):
        r=9+j; vals=[row.Canal,row.Cliente,int(row.Pedidos),row.Venta_Sol,row.Fac_Neto,row.Rechazado,row.Venta_Sol/tv]
        drow(ws,r,11,vals,alt=(r%2==0))
        for col in [14,15,16]: ws.cell(r,col).number_format='#,##0'
        ws.cell(r,17).number_format='0.0%'
    for i,w in enumerate([1,22,10,10,12,14,14,14,14,1,18,30,10,14,14,14,10],1):
        ws.column_dimensions[get_column_letter(i)].width=w
    print("  ✅ Dashboard")

    def sheet_canal(name, tab, title, mr, hdrs_list, data_df, data_fn, col_widths, n_tot_cols, tot_vals):
        ws_ = wb.create_sheet(name); ws_.sheet_view.showGridLines=False; ws_.sheet_properties.tabColor=tab
        title_row(ws_, title, mr)
        hdr(ws_, 2, range(1, len(hdrs_list)+1), hdrs_list)
        data_fn(ws_, data_df)
        tr_ = 3 + len(data_df)
        total_row(ws_, tr_, n_tot_cols, value_cols=tot_vals)
        for col, fmt in tot_vals.items():
            if isinstance(fmt, float) and fmt <= 1.0: ws_.cell(tr_,col).number_format='0.0%'
            elif isinstance(fmt, (int,float)): ws_.cell(tr_,col).number_format='#,##0'
        ws_.auto_filter.ref=f"A2:{get_column_letter(len(hdrs_list))}{tr_}"; ws_.freeze_panes='A3'
        for i,w in enumerate(col_widths,1): ws_.column_dimensions[get_column_letter(i)].width=w
        return ws_

    # ── Por Canal ────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("📦 Por Canal")
    ws4.sheet_view.showGridLines=False; ws4.sheet_properties.tabColor=C["GREEN"]
    title_row(ws4,"ANÁLISIS POR CANAL",'A1:J1')
    hdr(ws4,2,range(1,11),["Canal","Pedidos","Líneas","Unidades","Venta Sol. ($)","Fac. Bruto ($)","NC ZDEV ($)","Fac. Neto ($)","Rechazado ($)","Part. %"])
    for i,row in canal_df.iterrows():
        r=3+list(canal_df.index).index(i); nc_val=row.get('NC',0)
        vals=[row['Canal'],int(row['Pedidos']),int(row['Lineas']),int(row['Unidades']),
              row['Venta_Sol'],row['Fac_Bruto'],nc_val,row['Fac_Bruto']-nc_val,row['Rechazado'],row['Part_%']]
        drow(ws4,r,1,vals,alt=(r%2==0))
        for col in [5,6,7,8,9]: ws4.cell(r,col).number_format='#,##0'
        ws4.cell(r,10).number_format='0.0%'
        if nc_val>0: ws4.cell(r,7).fill=hfill(C["LRED"]); ws4.cell(r,7).font=Font(name="Arial",color=C["RED"])
        if row['Rechazado']>0: ws4.cell(r,9).fill=hfill(C["LORANG"])
    tr4=3+len(canal_df)
    total_row(ws4,tr4,10,value_cols={2:kpis['total_pedidos'],3:kpis['total_lineas'],5:tv,6:tfb,7:tnc,8:tfn,9:tr,10:1.0})
    for col in [5,6,7,8,9]: ws4.cell(tr4,col).number_format='#,##0'
    ws4.cell(tr4,10).number_format='0.0%'
    for w,col in zip([24,10,10,12,16,16,14,16,14,10],range(1,11)): ws4.column_dimensions[get_column_letter(col)].width=w
    chart=BarChart(); chart.type="bar"; chart.title="Venta por Canal ($)"; chart.style=10; chart.width=20; chart.height=12
    dr=Reference(ws4,min_col=5,max_col=5,min_row=2,max_row=2+len(canal_df))
    cats=Reference(ws4,min_col=1,min_row=3,max_row=2+len(canal_df))
    chart.add_data(dr,titles_from_data=True); chart.set_categories(cats); ws4.add_chart(chart,f"A{tr4+3}")
    ws4.auto_filter.ref=f"A2:J{tr4}"; ws4.freeze_panes='A3'
    print("  ✅ Por Canal")

    # ── Por Cliente ──────────────────────────────────────────────────────────
    ws5=wb.create_sheet("👥 Por Cliente"); ws5.sheet_view.showGridLines=False; ws5.sheet_properties.tabColor=C["AMBER"]
    title_row(ws5,f"TOP {CONFIG['top_clientes']} CLIENTES",'A1:J1')
    hdr(ws5,2,range(1,11),["#","Canal","Cliente","Vendedor Principal","Pedidos","Líneas","Venta Sol. ($)","Fac. Neto ($)","Rechazado ($)","Part. %"])
    for j,row in enumerate(cliente_df.itertuples(index=False),1):
        r=2+j
        nombre_vend = getattr(row, CONFIG["col_nombre_vend"].replace(' ','_').replace('.','_'), '') or ''
        vals=[j,row.Canal,row.Cliente,str(nombre_vend),int(row.Pedidos),int(row.Lineas),row.Venta_Sol,row.Fac_Neto,row.Rechazado,row.Venta_Sol/tv]
        drow(ws5,r,1,vals,alt=(r%2==0))
        ws5.cell(r,7).number_format='#,##0'; ws5.cell(r,8).number_format='#,##0'; ws5.cell(r,9).number_format='#,##0'
        ws5.cell(r,10).number_format='0.0%'
        if row.Rechazado>0: ws5.cell(r,9).fill=hfill(C["LORANG"])
    ws5.auto_filter.ref=f"A2:J{2+len(cliente_df)}"; ws5.freeze_panes='A3'
    for w,col in zip([5,20,35,24,10,10,16,16,14,10],range(1,11)): ws5.column_dimensions[get_column_letter(col)].width=w
    print("  ✅ Por Cliente")

    # ── Top SKUs ─────────────────────────────────────────────────────────────
    ws6=wb.create_sheet("🏷 Top SKUs"); ws6.sheet_view.showGridLines=False; ws6.sheet_properties.tabColor=C["RED"]
    title_row(ws6,f"TOP {CONFIG['top_skus']} SKUs",'A1:I1')
    hdr(ws6,2,range(1,10),["#","Canal","SKU SAP","Descripción","Unidades","Líneas","Clientes","Venta Sol. ($)","Rechazado ($)"])
    for j,row in enumerate(sku_df.itertuples(index=False),1):
        r=2+j; vals=[j,row.Canal,row.SKU_SAP,row.Descripcion,int(row.Unidades),int(row.Lineas),int(row.Clientes),row.Venta_Sol,row.Rechazado]
        drow(ws6,r,1,vals,alt=(r%2==0)); ws6.cell(r,8).number_format='#,##0'; ws6.cell(r,9).number_format='#,##0'
        if row.Rechazado>0: ws6.cell(r,9).fill=hfill(C["LORANG"])
    ws6.auto_filter.ref=f"A2:I{2+len(sku_df)}"; ws6.freeze_panes='A3'
    for w,col in zip([5,20,14,40,12,10,10,16,14],range(1,10)): ws6.column_dimensions[get_column_letter(col)].width=w
    print("  ✅ Top SKUs")

    # ── Evolución Diaria ─────────────────────────────────────────────────────
    ws7=wb.create_sheet("📈 Evolución Diaria"); ws7.sheet_view.showGridLines=False; ws7.sheet_properties.tabColor=C["BLUE"]
    title_row(ws7,"EVOLUCIÓN DIARIA",'A1:G1')
    hdr(ws7,2,range(1,8),["Canal","Fecha","Pedidos","Líneas","Venta Sol. ($)","Rechazado ($)","Venta Prom/Pedido ($)"])
    for j,row in enumerate(fecha_df.itertuples(index=False),1):
        r=2+j; prom=row.Venta_Sol/row.Pedidos if row.Pedidos else 0
        vals=[row.Canal,row.Fecha_str,int(row.Pedidos),int(row.Lineas),row.Venta_Sol,row.Rechazado,prom]
        drow(ws7,r,1,vals,alt=(r%2==0)); ws7.cell(r,5).number_format='#,##0'; ws7.cell(r,6).number_format='#,##0'; ws7.cell(r,7).number_format='#,##0'
        if row.Rechazado>0: ws7.cell(r,6).fill=hfill(C["LORANG"])
    ws7.conditional_formatting.add(f"E3:E{2+len(fecha_df)}",
        ColorScaleRule(start_type='min',start_color='FEE2E2',mid_type='percentile',mid_value=50,mid_color='FEF3C7',end_type='max',end_color='DCFCE7'))
    ws7.auto_filter.ref=f"A2:G{2+len(fecha_df)}"; ws7.freeze_panes='A3'
    for w,col in zip([20,14,10,10,18,14,18],range(1,8)): ws7.column_dimensions[get_column_letter(col)].width=w
    chart2=LineChart(); chart2.title="Venta Diaria ($)"; chart2.style=10; chart2.width=24; chart2.height=14
    dr2=Reference(ws7,min_col=5,max_col=5,min_row=2,max_row=2+len(fecha_df))
    cats2=Reference(ws7,min_col=2,min_row=3,max_row=2+len(fecha_df))
    chart2.add_data(dr2,titles_from_data=True); chart2.set_categories(cats2); ws7.add_chart(chart2,"I2")
    print("  ✅ Evolución Diaria")

    # ── Vendedores ────────────────────────────────────────────────────────────
    ws8=wb.create_sheet("🧑‍💼 Vendedores"); ws8.sheet_view.showGridLines=False; ws8.sheet_properties.tabColor=C["DGRAY"]
    title_row(ws8,"RENDIMIENTO POR VENDEDOR",'A1:N1')
    hdr(ws8,2,range(1,15),["#","Canal","Cód. Vendedor","Nombre Vendedor","Pedidos","Líneas","Clientes",
                            "Venta Sol. ($)","Fac. Bruto ($)","NC ZDEV ($)","Fac. Neto ($)","Rechazado ($)",
                            "Sin Entrega (Ped.)","Sin Entrega ($)"])
    for j,rdf in vend_df.iterrows():
        r=3+list(vend_df.index).index(j); rank=list(vend_df.index).index(j)+1
        vals=[rank,rdf['Canal'],str(rdf[CONFIG["col_vendedor"]]),str(rdf[CONFIG["col_nombre_vend"]]),
              int(rdf['Pedidos']),int(rdf['Lineas']),int(rdf['Clientes']),
              rdf['Venta_Sol'],rdf['Fac_Bruto'],rdf['NC_ZDEV'],rdf['Fac_Neto'],rdf['Rechazado'],
              int(rdf['SE_Ped']),rdf['SE_Val']]
        drow(ws8,r,1,vals,alt=(r%2==0))
        for col in [8,9,10,11,12,14]: ws8.cell(r,col).number_format='#,##0'
        ws8.cell(r,1).alignment=Alignment(horizontal='center',vertical='center')
        if rdf['NC_ZDEV']>0: ws8.cell(r,10).fill=hfill(C["LRED"]); ws8.cell(r,10).font=Font(name="Arial",color=C["RED"])
        if rdf['Rechazado']>0: ws8.cell(r,12).fill=hfill(C["LORANG"])
        if rdf['SE_Ped']>0: ws8.cell(r,13).fill=hfill(C["LAMBER"]); ws8.cell(r,13).font=Font(name="Arial",bold=True,color=C["AMBER"]); ws8.cell(r,14).fill=hfill(C["LAMBER"])
    tr8=3+len(vend_df)
    total_row(ws8,tr8,14,value_cols={8:tv,9:tfb,10:tnc,11:tfn,12:tr,13:kpis['total_se_ped'],14:kpis['total_se_val']})
    for col in [8,9,10,11,12,14]: ws8.cell(tr8,col).number_format='#,##0'
    ws8.conditional_formatting.add(f"H3:H{tr8-1}",
        ColorScaleRule(start_type='min',start_color='FEE2E2',mid_type='percentile',mid_value=50,mid_color='FEF3C7',end_type='max',end_color='DCFCE7'))
    ws8.auto_filter.ref=f"A2:N{tr8}"; ws8.freeze_panes='A3'
    for i,w in enumerate([5,22,14,28,10,10,10,16,16,14,16,14,16,16],1): ws8.column_dimensions[get_column_letter(i)].width=w
    print("  ✅ Vendedores")

    # ── Sin Entrega ───────────────────────────────────────────────────────────
    ws2=wb.create_sheet("🚨 Sin Entrega"); ws2.sheet_view.showGridLines=False; ws2.sheet_properties.tabColor=C["RED"]
    title_row(ws2,"CLIENTES SIN ENTREGA — PEDIDOS PENDIENTES",'A1:J1')
    ws2.merge_cells('A2:J2'); ws2['A2'].fill=hfill(C["RED"]); ws2.row_dimensions[2].height=4
    kpis_se=[("PEDIDOS SIN ENTREGA",str(kpis['total_se_ped']),C["RED"],C["LRED"]),
             ("CLIENTES AFECTADOS",str(base[base['Sin_Entrega']]['Cliente'].nunique()),C["ORANGE"],C["LORANG"]),
             ("VENTA EN RIESGO",f"${kpis['total_se_val']/1e6:.1f}M",C["RED"],C["LRED"]),
             ("CON BLOQUEO",str(int((base[base['Sin_Entrega']][CONFIG["col_bloq"]].notna()&(base[base['Sin_Entrega']][CONFIG["col_bloq"]].str.strip()!='')).sum())),C["ORANGE"],C["LORANG"])]
    for i,(title,val,color,bg) in enumerate(kpis_se):
        col=i*2+1
        ws2.merge_cells(start_row=3,start_column=col,end_row=3,end_column=col+1)
        c=ws2.cell(3,col,title); c.font=Font(name="Arial",bold=True,size=9,color=color); c.fill=hfill(bg); c.alignment=Alignment(horizontal='center',vertical='center')
        ws2.merge_cells(start_row=4,start_column=col,end_row=5,end_column=col+1)
        c=ws2.cell(4,col,val); c.font=Font(name="Arial",bold=True,size=17,color=color); c.fill=hfill(bg); c.alignment=Alignment(horizontal='center',vertical='center')
        ws2.merge_cells(start_row=6,start_column=col,end_row=6,end_column=col+1); ws2.cell(6,col).fill=hfill(color)
    for r in [3,4,5,6]: ws2.row_dimensions[r].height=20
    ws2.row_dimensions[4].height=28
    ws2.merge_cells('A8:J8'); c=ws2['A8']
    c.value="  RESUMEN POR CLIENTE — CON VENDEDOR RESPONSABLE"
    c.font=Font(name="Arial",bold=True,size=11,color=C["WHITE"]); c.fill=hfill(C["RED"]); c.alignment=Alignment(horizontal='left',vertical='center'); ws2.row_dimensions[8].height=22
    SE_MAX_DOCS = int(se_res['Pedidos'].max()) if len(se_res) else 1
    base_hdrs = ["#","Canal","Cliente","Cód. Vendedor","Nombre Vendedor","Pedidos","Líneas","Venta Sol. ($)","Bloqueados","Motivo Bloqueo"]
    doc_hdrs = []
    for d in range(1, SE_MAX_DOCS+1):
        doc_hdrs += [f"Doc {d}", f"Fecha {d}"]
    all_hdrs = base_hdrs + doc_hdrs
    n_se_cols = len(all_hdrs)
    hdr(ws2,9,range(1,n_se_cols+1),all_hdrs,C["ORANGE"])
    for j,rdf in se_res.iterrows():
        r=10+list(se_res.index).index(j); rank=list(se_res.index).index(j)+1
        pairs = rdf['DocFechaPairs'] if isinstance(rdf['DocFechaPairs'], list) else []
        doc_vals = []
        for d in range(SE_MAX_DOCS):
            if d < len(pairs):
                doc_vals += [pairs[d][0], pairs[d][1]]
            else:
                doc_vals += ['', '']
        vals=[rank,rdf['Canal'],rdf['Cliente'],rdf[CONFIG["col_vendedor"]],rdf[CONFIG["col_nombre_vend"]],
              int(rdf['Pedidos']),int(rdf['Lineas']),rdf['Venta_Sol'],int(rdf['Bloq']),rdf['Motivo']] + doc_vals
        drow(ws2,r,1,vals,alt=(r%2==0)); ws2.cell(r,1).alignment=Alignment(horizontal='center'); ws2.cell(r,8).number_format='#,##0'
        if rdf['Bloq']>0: ws2.cell(r,9).fill=hfill(C["LRED"]); ws2.cell(r,9).font=Font(name="Arial",bold=True,color=C["RED"]); ws2.cell(r,10).fill=hfill(C["LRED"])
        # Center fecha columns and highlight doc cols lightly
        for d in range(SE_MAX_DOCS):
            doc_col = 11 + d*2
            fecha_col = 12 + d*2
            if doc_col <= n_se_cols:
                ws2.cell(r,doc_col).alignment=Alignment(horizontal='center',vertical='center')
                ws2.cell(r,fecha_col).alignment=Alignment(horizontal='center',vertical='center')
    tr2=10+len(se_res)
    total_row(ws2,tr2,n_se_cols,value_cols={6:int(se_res['Pedidos'].sum()),7:int(se_res['Lineas'].sum()),8:se_res['Venta_Sol'].sum()})
    ws2.cell(tr2,8).number_format='#,##0'
    ws2.auto_filter.ref=f"A9:{get_column_letter(n_se_cols)}{9+len(se_res)}"; ws2.freeze_panes='A10'
    sep=tr2+2; ws2.merge_cells(start_row=sep,start_column=1,end_row=sep,end_column=14)
    c=ws2.cell(sep,1,"  DETALLE LÍNEAS SIN ENTREGA"); c.font=Font(name="Arial",bold=True,size=11,color=C["WHITE"])
    c.fill=hfill(C["RED"]); c.alignment=Alignment(horizontal='left',vertical='center'); ws2.row_dimensions[sep].height=22
    hdr(ws2,sep+1,range(1,15),["Canal","Doc. Vta.","Cliente","Localidad","Fecha","SKU SAP","Descripción","Cant.","Venta Sol. ($)","Est. SM","Bloq.","Motivo Bloq.","Vendedor","Nombre Vendedor"],C["ORANGE"])
    for j2,rdet in enumerate(se_det.itertuples(index=False),1):
        r=sep+1+j2; vals=list(rdet); drow(ws2,r,1,vals,alt=(r%2==0)); ws2.cell(r,9).number_format='#,##0'
    ws2.auto_filter.ref=f"A{sep+1}:N{sep+1+len(se_det)}"
    base_widths = [5,18,30,14,24,10,10,14,10,22]
    doc_widths = [14,13] * SE_MAX_DOCS
    for i,w in enumerate(base_widths + doc_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width=w
    print("  ✅ Sin Entrega")

    # ── Localidad ─────────────────────────────────────────────────────────────
    ws9=wb.create_sheet("📍 Por Localidad"); ws9.sheet_view.showGridLines=False
    title_row(ws9,f"TOP {CONFIG['top_localidades']} LOCALIDADES",'A1:F1')
    hdr(ws9,2,range(1,7),["#","Canal","Localidad","Pedidos","Clientes","Venta Sol. ($)"])
    for j,row in enumerate(loc_df.itertuples(index=False),1):
        r=2+j; vals=[j,row.Canal,row.Localidad,int(row.Pedidos),int(row.Clientes),row.Venta_Sol]
        drow(ws9,r,1,vals,alt=(r%2==0)); ws9.cell(r,6).number_format='#,##0'
    ws9.conditional_formatting.add(f"F3:F{2+len(loc_df)}",
        ColorScaleRule(start_type='min',start_color='FEE2E2',mid_type='percentile',mid_value=50,mid_color='FEF3C7',end_type='max',end_color='DCFCE7'))
    ws9.auto_filter.ref=f"A2:F{2+len(loc_df)}"; ws9.freeze_panes='A3'
    for w,col in zip([5,20,22,12,12,18],range(1,7)): ws9.column_dimensions[get_column_letter(col)].width=w
    print("  ✅ Por Localidad")

    # ── Notas Crédito ZDEV ────────────────────────────────────────────────────
    ws_nc=wb.create_sheet("🔴 Notas Crédito ZDEV"); ws_nc.sheet_view.showGridLines=False; ws_nc.sheet_properties.tabColor=C["RED"]
    title_row(ws_nc,"NOTAS DE CRÉDITO (ZDEV) — RESTADAS DEL FACTURADO",'A1:J1')
    ws_nc.merge_cells('A3:F3'); c=ws_nc['A3']
    c.value=f"  Total NC ZDEV: -${tnc:,.0f} CLP  (restado del facturado bruto)"
    c.font=Font(name="Arial",bold=True,size=11,color=C["WHITE"]); c.fill=hfill(C["RED"])
    c.alignment=Alignment(horizontal='left',vertical='center'); ws_nc.row_dimensions[3].height=24
    hdr(ws_nc,5,range(1,11),["Canal","Cliente","Doc. Vta.","Fecha","SKU SAP","Descripción","Cant.","Monto NC ($)","Vendedor","Nombre Vendedor"],C["RED"])
    for j,row in enumerate(nc_det.itertuples(index=False),1):
        r=5+j; vals=list(row); drow(ws_nc,r,1,vals,alt=(r%2==0)); ws_nc.cell(r,8).number_format='#,##0'
        ws_nc.cell(r,8).fill=hfill(C["LRED"]); ws_nc.cell(r,8).font=Font(name="Arial",color=C["RED"])
    ws_nc.auto_filter.ref=f"A5:J{5+len(nc_det)}"; ws_nc.freeze_panes='A6'
    for w,col in zip([18,30,12,12,12,38,10,14,12,24],range(1,11)): ws_nc.column_dimensions[get_column_letter(col)].width=w
    print("  ✅ Notas Crédito ZDEV")

    # Reordenar pestañas
    # ── Con Entrega No Facturado ─────────────────────────────────────────────
    if cenf_res is not None and len(cenf_res) > 0:
        ws_cf = wb.create_sheet("📬 Entrega No Facturado")
        ws_cf.sheet_view.showGridLines = False
        ws_cf.sheet_properties.tabColor = C["PURPLE"] if "PURPLE" in C else "7C3AED"

        # Title
        ws_cf.merge_cells('A1:N1')
        c = ws_cf['A1']
        c.value = f"  PEDIDOS CON ENTREGA Y SIN FACTURAR  |  {fecha_archivo}"
        c.font = Font(name="Arial", bold=True, size=13, color=C["WHITE"])
        c.fill = hfill("7C3AED"); c.alignment = Alignment(horizontal='left', vertical='center')
        ws_cf.row_dimensions[1].height = 28

        # KPI strip
        kpi_cf = [
            ("PEDIDOS SIN FACTURAR", str(kpis['cenf_docs']),          "7C3AED", "EDE9FE"),
            ("VENTA SOLICITADA",     f"${kpis['cenf_venta']/1e6:.1f}M","16A34A", "DCFCE7"),
            ("CANALES AFECTADOS",    str(kpis['cenf_canales']),        "D97706", "FEF3C7"),
            ("DIFERENCIA TOTAL",     f"${cenf_res['Diferencia'].sum()/1e6:.1f}M", "DC2626", "FEE2E2"),
        ]
        for i,(title,val,color,bg) in enumerate(kpi_cf):
            col=i*3+1
            ws_cf.merge_cells(start_row=3,start_column=col,end_row=3,end_column=col+2)
            c=ws_cf.cell(3,col,title); c.font=Font(name="Arial",bold=True,size=9,color=color)
            c.fill=hfill(bg); c.alignment=Alignment(horizontal='center',vertical='center')
            ws_cf.merge_cells(start_row=4,start_column=col,end_row=5,end_column=col+2)
            c=ws_cf.cell(4,col,val); c.font=Font(name="Arial",bold=True,size=17,color=color)
            c.fill=hfill(bg); c.alignment=Alignment(horizontal='center',vertical='center')
            ws_cf.merge_cells(start_row=6,start_column=col,end_row=6,end_column=col+2)
            ws_cf.cell(6,col).fill=hfill(color)
        for r in [3,4,5,6]: ws_cf.row_dimensions[r].height=20
        ws_cf.row_dimensions[4].height=28

        # Subtitle resumen
        ws_cf.merge_cells('A8:N8')
        c=ws_cf['A8']; c.value="  RESUMEN POR PEDIDO — CON ENTREGA GENERADA PERO AÚN NO FACTURADO"
        c.font=Font(name="Arial",bold=True,size=11,color=C["WHITE"])
        c.fill=hfill("7C3AED"); c.alignment=Alignment(horizontal='left',vertical='center')
        ws_cf.row_dimensions[8].height=22

        hdrs_cf = ["#","Canal","Doc. Vta.","N° Entrega","Cliente","Nombre Vendedor",
                   "Líneas","Venta Sol. ($)","Facturado ($)","Diferencia ($)","Rechazado ($)"]
        hdr(ws_cf, 9, range(1,12), hdrs_cf, "7C3AED")

        for j, rdf in cenf_res.iterrows():
            r = 10 + list(cenf_res.index).index(j)
            rank = list(cenf_res.index).index(j) + 1
            vals = [rank, rdf['Canal'], rdf[CONFIG["col_doc"]], rdf[CONFIG["col_entrega"]],
                    rdf['Cliente'], rdf[CONFIG["col_nombre_vend"]],
                    int(rdf['Lineas']), rdf['Venta_Sol'], rdf['Fac'],
                    rdf['Diferencia'], rdf['Rechazado']]
            drow(ws_cf, r, 1, vals, alt=(r%2==0))
            ws_cf.cell(r,1).alignment = Alignment(horizontal='center',vertical='center')
            for col in [8,9,10,11]: ws_cf.cell(r,col).number_format='#,##0'
            # Highlight diferencia en amarillo si > 0
            if rdf['Diferencia'] > 0:
                ws_cf.cell(r,10).fill = hfill(C["LAMBER"])
                ws_cf.cell(r,10).font = Font(name="Arial", bold=True, color=C["AMBER"])
            if rdf['Rechazado'] > 0:
                ws_cf.cell(r,11).fill = hfill(C["LORANG"])

        tr_cf = 10 + len(cenf_res)
        total_row(ws_cf, tr_cf, 11, value_cols={
            7: int(cenf_res['Lineas'].sum()),
            8: cenf_res['Venta_Sol'].sum(),
            9: cenf_res['Fac'].sum(),
            10: cenf_res['Diferencia'].sum(),
            11: cenf_res['Rechazado'].sum()})
        for col in [8,9,10,11]: ws_cf.cell(tr_cf,col).number_format='#,##0'

        ws_cf.auto_filter.ref = f"A9:K{9+len(cenf_res)}"
        ws_cf.freeze_panes = 'A10'

        # Detalle
        sep_cf = tr_cf + 2
        ws_cf.merge_cells(start_row=sep_cf,start_column=1,end_row=sep_cf,end_column=14)
        c=ws_cf.cell(sep_cf,1,"  DETALLE DE LÍNEAS CON ENTREGA Y SIN FACTURAR")
        c.font=Font(name="Arial",bold=True,size=11,color=C["WHITE"])
        c.fill=hfill("7C3AED"); c.alignment=Alignment(horizontal='left',vertical='center')
        ws_cf.row_dimensions[sep_cf].height=22

        det_hdrs_cf = ["Canal","Doc. Vta.","N° Entrega","Cliente","Localidad","Fecha",
                       "SKU SAP","Descripción","Cant.","Venta Sol. ($)","Facturado ($)","Rechazado ($)",
                       "Vendedor","Nombre Vendedor"]
        hdr(ws_cf, sep_cf+1, range(1,15), det_hdrs_cf, "7C3AED")
        for j2, rdet in enumerate(cenf_det.itertuples(index=False),1):
            r=sep_cf+1+j2; vals=list(rdet)
            drow(ws_cf,r,1,vals,alt=(r%2==0))
            ws_cf.cell(r,10).number_format='#,##0'
            ws_cf.cell(r,11).number_format='#,##0'
            ws_cf.cell(r,12).number_format='#,##0'
            if (vals[10] or 0) == 0 and (vals[9] or 0) > 0:
                ws_cf.cell(r,10).fill = hfill(C["LAMBER"])

        ws_cf.auto_filter.ref=f"A{sep_cf+1}:N{sep_cf+1+len(cenf_det)}"

        for i,w in enumerate([5,18,13,13,30,24,10,10,16,16,14,14,12,24],1):
            ws_cf.column_dimensions[get_column_letter(i)].width=w

        print("  ✅ Entrega No Facturado")

    order=["📊 Dashboard","🚨 Sin Entrega","📬 Entrega No Facturado","📋 Datos Limpios","📦 Por Canal",
           "👥 Por Cliente","🏷 Top SKUs","📈 Evolución Diaria","🧑‍💼 Vendedores",
           "📍 Por Localidad","🔴 Notas Crédito ZDEV"]
    for i,name in enumerate(order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i-wb.sheetnames.index(name))

    wb.save(ruta_salida)

# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("❌ Uso: python procesar_pedidos.py ARCHIVO.xls")
        print("   Ejemplo: python procesar_pedidos.py Pedidos_01-05-2026.xls")
        sys.exit(1)

    ruta_entrada = sys.argv[1]
    if not os.path.exists(ruta_entrada):
        print(f"❌ No se encontró el archivo: {ruta_entrada}")
        sys.exit(1)

    hoy = datetime.today().strftime("%Y-%m-%d")
    nombre_base = os.path.splitext(os.path.basename(ruta_entrada))[0]
    ruta_salida = f"Reporte_{nombre_base}_{hoy}.xlsx"

    print(f"\n{'='*55}")
    print(f"  PROCESANDO: {os.path.basename(ruta_entrada)}")
    print(f"{'='*55}")

    df, fecha_archivo = leer_archivo(ruta_entrada)
    df = limpiar(df)
    kpis, canal_df, cliente_df, fecha_df, sku_df, loc_df, vend_df, se_res, se_det, nc_det, base, sin_e, cenf_res, cenf_det = calcular_pivots(df)

    print(f"\n📊 Resumen:")
    print(f"   Pedidos:      {kpis['total_pedidos']:,}")
    print(f"   Venta Sol.:   ${kpis['total_venta']:,.0f}")
    print(f"   Fac. Neto:    ${kpis['total_fac_neto']:,.0f}")
    print(f"   NC ZDEV:      -${kpis['total_nc']:,.0f}")
    print(f"   Rechazado:    ${kpis['total_rechaz']:,.0f}")
    print(f"   Sin entrega:  {kpis['total_se_ped']} pedidos")
    print(f"   Tasa WMS:     {kpis['tasa_sm']:.1f}%")

    print(f"\n📝 Generando hojas...")
    escribir_excel(df, fecha_archivo, kpis, canal_df, cliente_df, fecha_df,
                   sku_df, loc_df, vend_df, se_res, se_det, nc_det, base, ruta_salida,
                   cenf_res=cenf_res, cenf_det=cenf_det)

    print(f"\n✅ Archivo generado:")
    print(f"   {ruta_salida}")
    print(f"{'='*55}\n")
